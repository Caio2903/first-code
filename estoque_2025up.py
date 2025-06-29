import sqlite3
from datetime import datetime
from tkinter import *
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
import os

conn = sqlite3.connect("produtos.db")
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS produtos (
    nome TEXT PRIMARY KEY,
    preco REAL NOT NULL,
    estoque INTEGER NOT NULL
)
""")
cursor.execute("""
CREATE TABLE IF NOT EXISTS pedidos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    produto TEXT NOT NULL,
    quantidade INTEGER NOT NULL,
    total REAL NOT NULL,
    data TEXT NOT NULL
)
""")
conn.commit()

def exportar_para_excel(produto, qtd, total):
    nome_arquivo = "pedidos.xlsx"
    if not os.path.exists(nome_arquivo):
        wb = Workbook()
        ws = wb.active
        ws.append(["Data", "Produto", "Quantidade", "Total"])
        wb.save(nome_arquivo)

    wb = load_workbook(nome_arquivo)
    ws = wb.active
    data = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([data, produto, qtd, total])
    wb.save(nome_arquivo)

def exportar_produtos_excel():
    nome_arquivo = "produtos.xlsx"
    if not os.path.exists(nome_arquivo):
        wb = Workbook()
        ws = wb.active
        ws.append(["Nome", "Preço", "Estoque"])
    else:
        wb = load_workbook(nome_arquivo)
        ws = wb.active
        ws.delete_rows(2, ws.max_row)  # limpa produtos antigos mantendo cabeçalho

    cursor.execute("SELECT * FROM produtos")
    for nome, preco, estoque in cursor.fetchall():
        ws.append([nome, preco, estoque])
    wb.save(nome_arquivo)
    messagebox.showinfo("Exportado", f"Produtos exportados para {nome_arquivo}.")

def cadastrar_produto():
    nome = entry_nome.get().strip()
    try:
        preco = float(entry_preco.get())
        estoque = int(entry_estoque.get())
    except:
        messagebox.showerror("Erro", "Preço ou estoque inválido.")
        return
    try:
        cursor.execute("INSERT INTO produtos VALUES (?, ?, ?)", (nome, preco, estoque))
        conn.commit()
        messagebox.showinfo("Sucesso", "Produto cadastrado.")
        atualizar_lista()
        atualizar_combo()
    except sqlite3.IntegrityError:
        messagebox.showwarning("Aviso", "Produto já existe.")

def editar_produto():
    nome = entry_nome.get().strip()
    try:
        preco = float(entry_preco.get())
        estoque = int(entry_estoque.get())
    except:
        messagebox.showerror("Erro", "Preço ou estoque inválido.")
        return
    cursor.execute("UPDATE produtos SET preco=?, estoque=? WHERE nome=?", (preco, estoque, nome))
    conn.commit()
    messagebox.showinfo("Editado", "Produto atualizado.")
    atualizar_lista()
    atualizar_combo()

def excluir_produto():
    nome = entry_nome.get().strip()
    cursor.execute("DELETE FROM produtos WHERE nome=?", (nome,))
    conn.commit()
    messagebox.showinfo("Deletado", "Produto removido.")
    atualizar_lista()
    atualizar_combo()

def atualizar_lista():
    for row in tree.get_children():
        tree.delete(row)
    cursor.execute("SELECT * FROM produtos")
    for produto in cursor.fetchall():
        tree.insert("", END, values=produto)

def preencher_campos(event):
    item = tree.focus()
    if not item:
        return
    valores = tree.item(item, "values")
    entry_nome.delete(0, END)
    entry_preco.delete(0, END)
    entry_estoque.delete(0, END)
    entry_nome.insert(0, valores[0])
    entry_preco.insert(0, valores[1])
    entry_estoque.insert(0, valores[2])

def fazer_pedido():
    produto = combo_produto.get()
    try:
        qtd = int(entry_qtd.get())
    except:
        messagebox.showerror("Erro", "Quantidade inválida.")
        return
    cursor.execute("SELECT preco, estoque FROM produtos WHERE nome=?", (produto,))
    res = cursor.fetchone()
    if not res:
        messagebox.showerror("Erro", "Produto não encontrado.")
        return
    preco, estoque = res
    if qtd > estoque:
        messagebox.showerror("Erro", "Estoque insuficiente.")
        return
    total = preco * qtd
    data = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute("INSERT INTO pedidos (produto, quantidade, total, data) VALUES (?, ?, ?, ?)",
                   (produto, qtd, total, data))
    cursor.execute("UPDATE produtos SET estoque = estoque - ? WHERE nome=?", (qtd, produto))
    conn.commit()
    exportar_para_excel(produto, qtd, total)
    messagebox.showinfo("Pedido", f"Pedido de {qtd}x {produto} registrado.\nTotal: R${total:.2f}")
    atualizar_lista()
    atualizar_combo()

def ver_historico():
    top = Toplevel(root)
    top.title("Histórico de Pedidos")
    tree_hist = ttk.Treeview(top, columns=("Produto", "Qtd", "Total", "Data"), show="headings")
    tree_hist.heading("Produto", text="Produto")
    tree_hist.heading("Qtd", text="Qtd")
    tree_hist.heading("Total", text="Total")
    tree_hist.heading("Data", text="Data")
    tree_hist.pack(fill=BOTH, expand=True)
    cursor.execute("SELECT produto, quantidade, total, data FROM pedidos ORDER BY data DESC")
    for linha in cursor.fetchall():
        tree_hist.insert("", END, values=linha)

def total_vendas():
    cursor.execute("SELECT SUM(total) FROM pedidos")
    total = cursor.fetchone()[0]
    total = total if total else 0
    messagebox.showinfo("Total de Vendas", f"Total vendido: R${total:.2f}")

root = Tk()
root.title("Sistema de Cadastro e Pedidos")

frame_produto = LabelFrame(root, text="Cadastro de Produtos", padx=10, pady=10)
frame_produto.pack(fill="both", expand="yes", padx=10, pady=5)

Label(frame_produto, text="Nome").grid(row=0, column=0)
Label(frame_produto, text="Preço").grid(row=0, column=1)
Label(frame_produto, text="Estoque").grid(row=0, column=2)

entry_nome = Entry(frame_produto)
entry_preco = Entry(frame_produto)
entry_estoque = Entry(frame_produto)

entry_nome.grid(row=1, column=0)
entry_preco.grid(row=1, column=1)
entry_estoque.grid(row=1, column=2)

Button(frame_produto, text="Cadastrar", command=cadastrar_produto).grid(row=1, column=3)
Button(frame_produto, text="Editar", command=editar_produto).grid(row=1, column=4)
Button(frame_produto, text="Excluir", command=excluir_produto).grid(row=1, column=5)

tree = ttk.Treeview(root, columns=("Nome", "Preço", "Estoque"), show="headings")
tree.heading("Nome", text="Nome")
tree.heading("Preço", text="Preço")
tree.heading("Estoque", text="Estoque")
tree.bind("<<TreeviewSelect>>", preencher_campos)
tree.pack(padx=10, pady=5, fill="x")

frame_pedido = LabelFrame(root, text="Realizar Pedido", padx=10, pady=10)
frame_pedido.pack(fill="both", expand="yes", padx=10, pady=5)

Label(frame_pedido, text="Produto").grid(row=0, column=0)
Label(frame_pedido, text="Quantidade").grid(row=0, column=1)

combo_produto = ttk.Combobox(frame_pedido, values=[])
entry_qtd = Entry(frame_pedido)

combo_produto.grid(row=1, column=0)
entry_qtd.grid(row=1, column=1)

Button(frame_pedido, text="Fazer Pedido", command=fazer_pedido).grid(row=1, column=2)

frame_hist = Frame(root)
frame_hist.pack(pady=5)
Button(frame_hist, text="Ver Histórico de Pedidos", command=ver_historico).pack(side=LEFT, padx=5)
Button(frame_hist, text="Total de Vendas", command=total_vendas).pack(side=LEFT, padx=5)
Button(frame_hist, text="Exportar Produtos p/ Excel", command=exportar_produtos_excel).pack(side=LEFT, padx=5)

label_assinatura = Label(root, text="Criado por @caio_h44", font=("Arial", 8), fg="gray")
label_assinatura.pack(side=BOTTOM, pady=5)

def atualizar_combo():
    cursor.execute("SELECT nome FROM produtos")
    nomes = [row[0] for row in cursor.fetchall()]
    combo_produto["values"] = nomes

atualizar_lista()
atualizar_combo()
root.mainloop()
conn.close()
